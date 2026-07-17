from django.db.models import Q

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied

from autenticacion.models import CorreoAutorizado
from ..models import FUS
from .helpers import _rol, ROLES_PARTICULAR, _propietario_fus


# ── Exportar FUS ──────────────────────────────────────────────────────────────

def _fus_queryset(request):
    rol = _rol(request.user)
    if rol not in ROLES_PARTICULAR:
        raise PermissionDenied('No autorizado.')
    qs = FUS.objects.filter(activo=1).select_related(
        'idSolicitanteInterno', 'idMedioRecepcion'
    ).order_by('-fechaRegistro')
    if rol == 'EQUIPO_PARTICULAR':
        propietario = _propietario_fus(request.user)
        qs = qs.filter(idSolicitanteInterno=propietario) if propietario else qs.none()
    estatus = request.query_params.get('estatusParticular')
    search  = request.query_params.get('search')
    if estatus: qs = qs.filter(estatusParticular_id=estatus)
    if search:
        emails_nombre = list(CorreoAutorizado.objects.filter(nombre__icontains=search).values_list('email', flat=True))
        qs = qs.filter(
            Q(folio__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(contexto__icontains=search) |
            Q(medioEspecificacion__icontains=search) |
            Q(criterios__icontains=search) |
            Q(nombreExterno__icontains=search) |
            Q(telefonoExterno__icontains=search) |
            Q(correoExterno__icontains=search) |
            Q(idMedioRecepcion__nombreMedio__icontains=search) |
            Q(idSolicitanteInterno__email__icontains=search) |
            Q(idSolicitanteInterno__email__in=emails_nombre) |
            Q(evidencias__nombreArchivo__icontains=search) |
            Q(evidencias__comentarios__icontains=search) |
            Q(turnados__solicitudTexto__icontains=search) |
            Q(turnados__idMedio__nombreMedio__icontains=search) |
            Q(turnados__idRemitente__email__icontains=search) |
            Q(turnados__idRemitente__email__in=emails_nombre) |
            Q(turnados__idDestinatario__email__icontains=search) |
            Q(turnados__idDestinatario__email__in=emails_nombre) |
            Q(turnados__seguimientos__descripcionActividad__icontains=search) |
            Q(turnados__seguimientos__accionTexto__icontains=search)
        ).distinct()
    return qs


class ExportarFUSExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse

        qs = _fus_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Solicitudes FUS'

        headers = ['Folio', 'Fecha', 'Solicitante', 'Medio', 'Prioridad',
                   'Estatus', 'Descripción', 'Contexto']
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill('solid', fgColor='9F2241')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        fmt = lambda d: d.strftime('%d/%m/%Y %H:%M') if d else ''
        fmt_date = lambda d: d.strftime('%d/%m/%Y') if d else ''

        for fus in qs:
            sol = fus.idSolicitanteInterno
            nombre_sol = f"{sol.first_name} {sol.last_name}".strip() if sol else ''
            ws.append([
                fus.folio,
                fmt(fus.fechaHora),
                nombre_sol or (sol.email if sol else ''),
                fus.idMedioRecepcion.nombreMedio if fus.idMedioRecepcion else '',
                fus.prioridad or '',
                fus.estatusParticular_id,
                fus.descripcion,
                fus.contexto or '',
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="solicitudes_fus.xlsx"'
        wb.save(response)
        return response


class ExportarFUSPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io

        qs = _fus_queryset(request)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                     fontSize=16, textColor=colors.HexColor('#9F2241'),
                                     spaceAfter=6)
        cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)

        elements = [
            Paragraph('Reporte de Solicitudes FUS — ANAM', title_style),
            Spacer(1, 0.3*cm),
        ]

        fmt = lambda d: d.strftime('%d/%m/%Y %H:%M') if d else '—'
        fmt_date = lambda d: d.strftime('%d/%m/%Y') if d else '—'

        data = [['Folio', 'Fecha', 'Solicitante', 'Prioridad', 'Estatus', 'Descripción']]
        for fus in qs:
            sol = fus.idSolicitanteInterno
            nombre_sol = f"{sol.first_name} {sol.last_name}".strip() if sol else (sol.email if sol else '—')
            desc = fus.descripcion[:120] + ('…' if len(fus.descripcion) > 120 else '')
            data.append([
                Paragraph(fus.folio, cell_style),
                Paragraph(fmt(fus.fechaHora), cell_style),
                Paragraph(nombre_sol, cell_style),
                Paragraph(fus.prioridad or '—', cell_style),
                Paragraph(fus.estatusParticular_id, cell_style),
                Paragraph(desc, cell_style),
            ])

        col_widths = [4*cm, 3.5*cm, 3.5*cm, 2*cm, 2.5*cm, None]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F4F6')]),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#DDD0D5')),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="solicitudes_fus.pdf"'
        return response
