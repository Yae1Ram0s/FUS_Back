import datetime
import os

from django.db.models import Q
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from autenticacion.models import CorreoAutorizado
from ..models import FUS, Bitacora
from ..helpers import _resolver_unidad_administrativa
from .helpers import _rol, ROL2_ACCIONES, _metadata_generacion


BITACORA_COLS_VALIDAS  = ['folio', 'nombre', 'usuario', 'fecha', 'accion', 'estado_ant', 'estado_nuevo', 'observaciones']
BITACORA_COLS_DEFAULT  = ['folio', 'nombre', 'usuario', 'fecha', 'accion']


def _parse_columnas_bitacora(request):
    raw = request.query_params.get('columnas')
    if not raw:
        return BITACORA_COLS_DEFAULT
    cols = [c.strip() for c in raw.split(',') if c.strip() in BITACORA_COLS_VALIDAS]
    return cols or BITACORA_COLS_DEFAULT


BITACORA_TITULO_PRINCIPAL = 'Sistema de control de solicitudes'
ANAM_LOGO_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
    'frontend', 'src', 'assets', 'Logos_P_Hacienda_ANAM.png',
)


def _bitacora_base_qs(request):
    """Devuelve el queryset de bitácora ya filtrado por rol del usuario.
    Solo se incluyen registros ligados a un FUS (se excluyen eventos de cuenta
    como inicio/cierre de sesión o restablecimiento de contraseña)."""
    rol = _rol(request.user)
    if rol == 'ROL1':
        qs = Bitacora.objects.all()
    elif rol == 'ROL2':
        qs = Bitacora.objects.filter(
            usuario=request.user.email, accion__in=ROL2_ACCIONES
        )
    else:
        return Bitacora.objects.none()
    return qs.exclude(fusFolio__isnull=True).exclude(fusFolio='')


def _parse_fecha_local(fecha_str, fin_de_dia=False):
    """Convierte 'YYYY-MM-DD' (hora local del servidor) a datetime aware en UTC.
    Se evita el lookup `__date` de Django porque requiere CONVERT_TZ en MySQL,
    lo cual falla silenciosamente (devuelve NULL) si el servidor no tiene
    cargadas las tablas de zonas horarias, vaciando cualquier filtro de fecha."""
    try:
        y, m, d = (int(p) for p in fecha_str.split('-'))
    except (ValueError, AttributeError):
        return None
    naive = datetime.datetime(y, m, d, 23, 59, 59, 999999) if fin_de_dia else datetime.datetime(y, m, d, 0, 0, 0)
    return timezone.make_aware(naive, timezone.get_current_timezone())


def _aplicar_filtros_bitacora(qs, params, rol):
    usuario     = params.get('usuario')
    accion      = params.get('accion')
    folio       = params.get('folio')
    nombre      = params.get('nombre')
    estatus_fus = params.get('estatus_fus')
    fecha_desde = params.get('fecha_desde')
    fecha_hasta = params.get('fecha_hasta')

    q = params.get('q')
    if q:
        if rol == 'ROL1':
            emails_nombre = CorreoAutorizado.objects.filter(nombre__icontains=q).values_list('email', flat=True)
            qs = qs.filter(Q(fusFolio__icontains=q) | Q(usuario__icontains=q) | Q(usuario__in=list(emails_nombre)))
        else:
            qs = qs.filter(fusFolio__icontains=q)

    if usuario and rol == 'ROL1': qs = qs.filter(usuario__icontains=usuario)
    if accion:      qs = qs.filter(accion=accion)
    if folio:       qs = qs.filter(fusFolio__icontains=folio)

    if nombre and rol == 'ROL1':
        emails = CorreoAutorizado.objects.filter(nombre__icontains=nombre).values_list('email', flat=True)
        qs = qs.filter(usuario__in=list(emails))

    if estatus_fus == 'Vencido':
        folios = FUS.objects.filter(estatusParticular_id='Turnado', fechaLimite__lt=timezone.now()).values_list('folio', flat=True)
        qs = qs.filter(fusFolio__in=list(folios))
    elif estatus_fus == 'PorVencer':
        ahora = timezone.now()
        folios = FUS.objects.filter(
            estatusParticular_id='Turnado',
            fechaLimite__gte=ahora,
            fechaLimite__lte=ahora + datetime.timedelta(hours=24),
        ).values_list('folio', flat=True)
        qs = qs.filter(fusFolio__in=list(folios))
    elif estatus_fus:
        folios = FUS.objects.filter(estatusParticular_id=estatus_fus).values_list('folio', flat=True)
        qs = qs.filter(fusFolio__in=list(folios))

    dt_desde = _parse_fecha_local(fecha_desde) if fecha_desde else None
    if dt_desde: qs = qs.filter(fechaHora__gte=dt_desde)

    dt_hasta = _parse_fecha_local(fecha_hasta, fin_de_dia=True) if fecha_hasta else None
    if dt_hasta: qs = qs.filter(fechaHora__lte=dt_hasta)

    return qs


class BitacoraListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rol = _rol(request.user)
        ordering = request.query_params.get('ordering')
        qs = _bitacora_base_qs(request)
        if not ordering:
            qs = qs.order_by('-fechaHora')
        qs  = _aplicar_filtros_bitacora(qs, request.query_params, rol)

        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(100, max(1, int(request.query_params.get('page_size', 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50

        ORDERING_MAP = {
            'folio': 'fusFolio', 'fecha': 'fechaHora', 'accion': 'accion',
        }
        if ordering:
            campo = ordering.lstrip('-')
            campo_real = ORDERING_MAP.get(campo)
            if campo_real:
                qs = qs.order_by(f"{'-' if ordering.startswith('-') else ''}{campo_real}")

        total  = qs.count()
        offset = (page - 1) * page_size

        from ..serializers import BitacoraSerializer
        pagina = list(qs[offset: offset + page_size])
        nombres_map = dict(
            CorreoAutorizado.objects.filter(
                email__in={r.usuario for r in pagina}
            ).values_list('email', 'nombre')
        )
        data = BitacoraSerializer(pagina, many=True, context={'nombres_map': nombres_map}).data
        return Response({
            'total': total, 'page': page, 'page_size': page_size,
            'results': data, 'rol': rol,
        })


# ── Exportar Bitácora ────────────────────────────────────────────────────────

class ExportarBitacoraExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage

        rol = _rol(request.user)
        qs  = _bitacora_base_qs(request).order_by('-fechaHora')
        qs  = _aplicar_filtros_bitacora(qs, request.query_params, rol)
        columnas = _parse_columnas_bitacora(request)

        ACCION_LABELS = {
            'REGISTRO_FUS': 'Registro FUS', 'TURNAR_FUS': 'Turnar FUS',
            'ASIGNACION_ESTADO': 'Cambio de estado', 'REGISTRO_RESPUESTA': 'Registro respuesta',
            'REGISTRO_ACCION': 'Registro acción', 'CONCLUSION_FUS': 'Conclusión FUS',
            'INICIO_SESION': 'Inicio sesión', 'CIERRE_SESION': 'Cierre sesión',
            'RESTABLECER_CONTRASENA': 'Restablecer contraseña',
            'ELIMINACION': 'Eliminación',
        }
        nombres_map = dict(
            CorreoAutorizado.objects.filter(
                email__in=qs.values_list('usuario', flat=True).distinct()
            ).values_list('email', 'nombre')
        )
        col_defs = {
            'folio':         ('Folio',           lambda r: r.fusFolio or ''),
            'nombre':        ('Nombre',          lambda r: nombres_map.get(r.usuario, '')),
            'usuario':       ('Usuario',         lambda r: r.usuario),
            'fecha':         ('Fecha y hora',    lambda r: r.fechaHora.strftime('%d/%m/%Y %H:%M:%S') if r.fechaHora else ''),
            'accion':        ('Acción',          lambda r: ACCION_LABELS.get(r.accion, r.accion)),
            'estado_ant':    ('Estado anterior', lambda r: r.estadoAnterior or ''),
            'estado_nuevo':  ('Estado nuevo',    lambda r: r.estadoNuevo or ''),
            'observaciones': ('Observaciones',   lambda r: r.observaciones or ''),
        }
        headers = [col_defs[c][0] for c in columnas]
        n_cols  = len(headers)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bitácora'

        unidad_admin    = _resolver_unidad_administrativa(request.user)
        metadata_linea  = _metadata_generacion()

        # Encabezado institucional
        ws.append(['AGENCIA NACIONAL DE ADUANAS DE MÉXICO'])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Título principal
        ws.append([BITACORA_TITULO_PRINCIPAL.upper()])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Subtítulo dinámico — unidad/aduana asignada al usuario
        ws.append([unidad_admin.upper()])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(bold=True, size=11, color='1F5647')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.append([])

        # Metadatos — fecha/hora de generación (derecha)
        ws.append([metadata_linea])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(italic=True, size=9, color='595959')
        cell.alignment = Alignment(horizontal='right')

        # Logo institucional (izquierda)
        logo_row = ws.max_row + 1
        for _ in range(5):
            ws.append([])
        if os.path.exists(ANAM_LOGO_PATH):
            img = XLImage(ANAM_LOGO_PATH)
            img.width, img.height = 460, 460 * (150/889)
            ws.add_image(img, f'A{logo_row}')

        ws.append([])

        ws.append(headers)
        header_row = ws.max_row
        fill   = PatternFill('solid', fgColor='FFFF00')
        thin   = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[header_row]:
            cell.fill = fill
            cell.font = Font(bold=True, color='000000', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for r in qs:
            ws.append([col_defs[c][1](r) for c in columnas])
            for cell in ws[ws.max_row]:
                cell.border = border

        for i in range(1, n_cols + 1):
            col_letter = get_column_letter(i)
            max_len = max(
                (len(str(ws.cell(row=r, column=i).value or '')) for r in range(header_row, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        ws.freeze_panes = f'A{header_row + 1}'

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename="bitacora.xlsx"'
        wb.save(resp)
        return resp


class ExportarBitacoraPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io

        rol = _rol(request.user)
        qs  = _bitacora_base_qs(request).order_by('-fechaHora')
        qs  = _aplicar_filtros_bitacora(qs, request.query_params, rol)
        columnas = _parse_columnas_bitacora(request)

        ACCION_LABELS = {
            'REGISTRO_FUS': 'Registro FUS', 'TURNAR_FUS': 'Turnar FUS',
            'ASIGNACION_ESTADO': 'Cambio de estado', 'REGISTRO_RESPUESTA': 'Registro respuesta',
            'REGISTRO_ACCION': 'Registro acción', 'CONCLUSION_FUS': 'Conclusión FUS',
            'INICIO_SESION': 'Inicio sesión', 'CIERRE_SESION': 'Cierre sesión',
            'RESTABLECER_CONTRASENA': 'Restablecer contraseña',
            'ELIMINACION': 'Eliminación',
        }
        nombres_map = dict(
            CorreoAutorizado.objects.filter(
                email__in=qs.values_list('usuario', flat=True).distinct()
            ).values_list('email', 'nombre')
        )

        styles = getSampleStyleSheet()
        cell_style  = ParagraphStyle('cell', parent=styles['Normal'], fontSize=7.5, leading=10)

        col_defs = {
            'folio':         ('Folio',         lambda r: r.fusFolio or '—'),
            'nombre':        ('Nombre',        lambda r: nombres_map.get(r.usuario, '—')),
            'usuario':       ('Usuario',       lambda r: r.usuario),
            'fecha':         ('Fecha y hora',  lambda r: r.fechaHora.strftime('%d/%m/%Y %H:%M') if r.fechaHora else '—'),
            'accion':        ('Acción',        lambda r: ACCION_LABELS.get(r.accion, r.accion)),
            'estado_ant':    ('Estado ant.',   lambda r: r.estadoAnterior or '—'),
            'estado_nuevo':  ('Estado nuevo',  lambda r: r.estadoNuevo or '—'),
            'observaciones': ('Observaciones', lambda r: r.observaciones or '—'),
        }
        headers = [col_defs[c][0] for c in columnas]

        unidad_admin   = _resolver_unidad_administrativa(request.user)
        metadata_linea = _metadata_generacion()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1.2*cm, rightMargin=1.2*cm,
                                topMargin=4.4*cm, bottomMargin=1.2*cm)

        ancho_total = landscape(A4)[0] - 2.4*cm

        def _encabezado_bitacora(canvas_, doc_):
            canvas_.saveState()
            page_w, page_h = landscape(A4)
            centro = page_w / 2
            derecha = page_w - 1.2*cm
            y = page_h - 1.1*cm

            canvas_.setFont('Helvetica-Bold', 11)
            canvas_.setFillColor(colors.black)
            canvas_.drawCentredString(centro, y, 'AGENCIA NACIONAL DE ADUANAS DE MÉXICO')
            y -= 0.42*cm
            canvas_.drawCentredString(centro, y, BITACORA_TITULO_PRINCIPAL.upper())
            y -= 0.42*cm
            canvas_.setFillColor(colors.HexColor('#1F5647'))
            canvas_.drawCentredString(centro, y, unidad_admin.upper())
            canvas_.setFillColor(colors.black)

            y -= 0.5*cm
            canvas_.setFont('Helvetica-Oblique', 7)
            canvas_.setFillColor(colors.HexColor('#595959'))
            canvas_.drawRightString(derecha, y, metadata_linea)
            canvas_.setFillColor(colors.black)

            if os.path.exists(ANAM_LOGO_PATH):
                logo_w = 8*cm
                logo_h = logo_w * (150/889)
                canvas_.drawImage(ANAM_LOGO_PATH, 1.2*cm, y - logo_h - 0.2*cm,
                                  width=logo_w, height=logo_h, mask='auto', preserveAspectRatio=True)
            canvas_.restoreState()

        elements = []

        data = [headers]
        for r in qs:
            data.append([Paragraph(str(col_defs[c][1](r)), cell_style) for c in columnas])

        col_width = ancho_total / len(columnas)
        t = Table(data, colWidths=[col_width] * len(columnas), repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFFF00')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.black),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 8.5),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)

        doc.build(elements, onFirstPage=_encabezado_bitacora, onLaterPages=_encabezado_bitacora)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="bitacora.pdf"'
        return resp
