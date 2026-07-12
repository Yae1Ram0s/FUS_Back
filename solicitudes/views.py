import datetime
import hashlib
import json
import os
import uuid

ALLOWED_MIME_TYPES = {
    'application/pdf',
    'image/jpeg',
    'image/png',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
}
ALLOWED_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png', '.docx'}
MAX_FILE_SIZE      = 10 * 1024 * 1024   # 10 MB por archivo
MAX_TOTAL_SIZE     = 30 * 1024 * 1024   # 30 MB por FUS


def _validar_archivo(archivo):
    ext  = os.path.splitext(archivo.name)[1].lower()
    mime = (archivo.content_type or '').split(';')[0].strip()
    if ext not in ALLOWED_EXTENSIONS:
        return f'Extensión no permitida: {ext}. Usa PDF, JPG, PNG o DOCX.'
    if mime and mime not in ALLOWED_MIME_TYPES:
        return f'Tipo de archivo no permitido: {mime}.'
    if archivo.size > MAX_FILE_SIZE:
        return f'"{archivo.name}" supera 10 MB.'
    return None


from django.contrib.auth.models import User
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.exceptions import PermissionDenied

from autenticacion.models import CorreoAutorizado
from catalogos.models import MedioRecepcion
from .models import FUS, Evidencia, Turnado, Seguimiento, Bitacora, Notificacion
from .serializers import FUSSerializer, TurnadoSerializer, TurnadoActividadSerializer, SeguimientoSerializer, NotificacionSerializer
from .utils import get_rol, log_bitacora, resolver_nombre


def _push_notificacion(notif):
    from asgiref.sync import async_to_sync
    from channels.layers import get_channel_layer
    channel_layer = get_channel_layer()
    if channel_layer is None:
        return
    group_name = f'notificaciones_{notif.idDestinatario_id}'
    data = {
        'id':            str(notif.id),
        'fusFolio':      notif.fusFolio,
        'tipo':          notif.tipoEvento,
        'mensaje':       notif.mensaje,
        'leida':         False,
        'fechaCreacion': notif.fechaGeneracion.isoformat(),
    }
    async_to_sync(channel_layer.group_send)(
        group_name,
        {'type': 'nueva_notificacion', 'data': data},
    )


# ── Helpers ─────────────────────────────────────────────────────────────────

# _rol/_log se mantienen como alias locales para no tocar cada llamada existente.
_rol = get_rol
_log = log_bitacora


_ROL_FOLIO = {'ROL1': 'PARTICULAR', 'ROL2': 'TITULAR'}

def _generar_folio(rol, year):
    from django.db import transaction
    with transaction.atomic():
        seq = FUS.objects.select_for_update().filter(fechaRegistro__year=year).count() + 1
        segmento = _ROL_FOLIO.get(rol, rol)
        return f"ANAM/{segmento}/FUS/{seq:04d}/{year}"


def _sha256(f):
    h = hashlib.sha256()
    for chunk in f.chunks():
        h.update(chunk)
    return h.hexdigest()


def _guardar_evidencias(fus, request, user):
    """Valida y guarda los archivos de 'evidencias' del request para un FUS.
    Devuelve una Response de error si algo falla, o None si todo salió bien."""
    from django.conf import settings

    archivos = request.FILES.getlist('evidencias')
    if not archivos:
        return None

    total_size = sum(a.size for a in archivos)
    if total_size > MAX_TOTAL_SIZE:
        return Response({'detail': 'El total de archivos supera 30 MB.'}, status=status.HTTP_400_BAD_REQUEST)
    for archivo in archivos:
        err = _validar_archivo(archivo)
        if err:
            return Response({'detail': err}, status=status.HTTP_400_BAD_REQUEST)

    try:
        comentarios_lista = json.loads(request.data.get('comentariosEvidencias') or '[]')
    except (ValueError, TypeError):
        comentarios_lista = []

    for i, archivo in enumerate(archivos):
        sha = _sha256(archivo)
        nombre_seguro = os.path.basename(archivo.name)
        nombre_fisico = f"{uuid.uuid4().hex}_{nombre_seguro}"
        ruta_rel = f"evidencias/{fus.pk}/{nombre_fisico}"
        ruta_abs = os.path.join(settings.MEDIA_ROOT, ruta_rel)
        os.makedirs(os.path.dirname(ruta_abs), exist_ok=True)
        with open(ruta_abs, 'wb') as dest:
            for chunk in archivo.chunks():
                dest.write(chunk)
        comentario = comentarios_lista[i].strip() if i < len(comentarios_lista) and comentarios_lista[i] else None
        Evidencia.objects.create(
            idFus=fus,
            nombreArchivo=nombre_seguro,
            rutaArchivo=ruta_rel,
            tipoMime=archivo.content_type,
            hashSha256=sha,
            comentarios=comentario,
            idUsuarioRegistra=user.id,
        )
    return None


# ── FUS ─────────────────────────────────────────────────────────────────────

class FUSListCreateView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]

    def get(self, request):
        if _rol(request.user) != 'ROL1':
            return Response({'detail': 'No autorizado.'}, status=403)

        qs = FUS.objects.filter(activo=1).select_related(
            'idSolicitanteInterno', 'idMedioRecepcion'
        ).prefetch_related('evidencias')

        estatus = request.query_params.get('estatusParticular')
        search  = request.query_params.get('search')
        if estatus:
            qs = qs.filter(estatusParticular_id=estatus)
        if search:
            emails_nombre = list(CorreoAutorizado.objects.filter(nombre__icontains=search).values_list('email', flat=True))
            qs = qs.filter(
                Q(folio__icontains=search) |
                Q(descripcion__icontains=search) |
                Q(contexto__icontains=search) |
                Q(medioEspecificacion__icontains=search) |
                Q(criterios__icontains=search) |
                Q(nombreExterno__icontains=search) |
                Q(telefonoExterno__icontains=search) |
                Q(correoExterno__icontains=search) |
                Q(idMedioRecepcion__nombreMedio__icontains=search) |
                Q(idSolicitanteInterno__email__icontains=search) |
                Q(idSolicitanteInterno__email__in=emails_nombre) |
                Q(evidencias__nombreArchivo__icontains=search) |
                Q(evidencias__comentarios__icontains=search) |
                Q(turnados__solicitudTexto__icontains=search) |
                Q(turnados__idMedio__nombreMedio__icontains=search) |
                Q(turnados__idRemitente__email__icontains=search) |
                Q(turnados__idRemitente__email__in=emails_nombre) |
                Q(turnados__idDestinatario__email__icontains=search) |
                Q(turnados__idDestinatario__email__in=emails_nombre) |
                Q(turnados__seguimientos__descripcionActividad__icontains=search) |
                Q(turnados__seguimientos__accionTexto__icontains=search)
            ).distinct()

        qs = qs.order_by('-fechaRegistro')

        # Paginación
        try:
            page     = max(1, int(request.query_params.get('page', 1)))
            page_size = min(100, max(1, int(request.query_params.get('page_size', 30))))
        except (ValueError, TypeError):
            page, page_size = 1, 30

        total  = qs.count()
        offset = (page - 1) * page_size
        data   = FUSSerializer(qs[offset: offset + page_size], many=True).data
        return Response({'total': total, 'page': page, 'page_size': page_size, 'results': data})

    def post(self, request):
        user  = request.user
        rol   = _rol(user)
        if rol != 'ROL1':
            return Response({'detail': 'No autorizado.'}, status=403)

        data  = request.data
        ip    = request.META.get('REMOTE_ADDR')
        now   = timezone.now()
        year  = now.year

        medio_id = data.get('idMedioRecepcion')
        medio    = get_object_or_404(MedioRecepcion, pk=medio_id) if medio_id else None

        nombre_ext = data.get('nombreExterno', '').strip() or None
        tel_ext    = data.get('telefonoExterno', '').strip() or None
        correo_ext = data.get('correoExterno', '').strip() or None

        from django.db import IntegrityError

        fus = None
        for intento in range(3):
            folio = _generar_folio(rol, year)
            try:
                fus = FUS.objects.create(
                    folio=folio,
                    idSolicitanteInterno=user,
                    fechaHora=now,
                    descripcion=data.get('descripcion', ''),
                    contexto=data.get('contexto', ''),
                    idMedioRecepcion=medio,
                    medioEspecificacion=data.get('medioEspecificacion', ''),
                    prioridad=data.get('prioridad') or None,
                    criterios=data.get('criterios') or None,
                    nombreExterno=nombre_ext,
                    telefonoExterno=tel_ext,
                    correoExterno=correo_ext,
                    estatusParticular_id='Registrado',
                    idUsuarioRegistra=user.id,
                )
                break
            except IntegrityError:
                if intento == 2:
                    raise
                continue

        err_resp = _guardar_evidencias(fus, request, user)
        if err_resp:
            fus.delete()
            return err_resp

        _log(usuario=user.email, rol=rol, accion='REGISTRO_FUS',
             ip=ip, folio=folio, estado_nuevo='Registrado')

        return Response(FUSSerializer(fus).data, status=status.HTTP_201_CREATED)


class FUSDetailView(APIView):
    """GET / PATCH — ver o editar un FUS individual (ROL1). Solo editable en estatus 'Registrado'."""
    permission_classes = [IsAuthenticated]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]

    def get(self, request, pk):
        if _rol(request.user) != 'ROL1':
            return Response({'detail': 'No autorizado.'}, status=403)
        fus = get_object_or_404(
            FUS.objects.select_related('idSolicitanteInterno', 'idMedioRecepcion').prefetch_related('evidencias'),
            pk=pk, activo=1, idSolicitanteInterno=request.user,
        )
        return Response(FUSSerializer(fus).data)

    def patch(self, request, pk):
        user = request.user
        rol  = _rol(user)
        if rol != 'ROL1':
            return Response({'detail': 'No autorizado.'}, status=403)

        fus = get_object_or_404(FUS, pk=pk, activo=1, idSolicitanteInterno=user)
        if fus.estatusParticular_id != 'Registrado':
            return Response(
                {'detail': 'Solo se puede editar una solicitud en estatus "Registrado".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = request.data
        if 'idMedioRecepcion' in data:
            medio_id = data.get('idMedioRecepcion')
            fus.idMedioRecepcion = get_object_or_404(MedioRecepcion, pk=medio_id) if medio_id else None
        if 'descripcion' in data:         fus.descripcion = data.get('descripcion', '')
        if 'contexto' in data:            fus.contexto = data.get('contexto', '')
        if 'medioEspecificacion' in data:  fus.medioEspecificacion = data.get('medioEspecificacion', '')
        if 'prioridad' in data:           fus.prioridad = data.get('prioridad') or None
        if 'criterios' in data:           fus.criterios = data.get('criterios') or None
        if 'nombreExterno' in data:       fus.nombreExterno = data.get('nombreExterno', '').strip() or None
        if 'telefonoExterno' in data:     fus.telefonoExterno = data.get('telefonoExterno', '').strip() or None
        if 'correoExterno' in data:       fus.correoExterno = data.get('correoExterno', '').strip() or None
        fus.idUsuarioModifica = user.id
        fus.save()

        err_resp = _guardar_evidencias(fus, request, user)
        if err_resp:
            return err_resp

        _log(usuario=user.email, rol=rol, accion='REGISTRO_FUS',
             ip=request.META.get('REMOTE_ADDR'), folio=fus.folio, obs='Edición de solicitud')

        return Response(FUSSerializer(fus).data)


class TurnarFUSView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        user = request.user
        rol  = _rol(user)
        if rol != 'ROL1':
            return Response({'detail': 'No autorizado.'}, status=403)

        fus           = get_object_or_404(FUS, pk=pk, activo=1, idSolicitanteInterno=user)
        ip            = request.META.get('REMOTE_ADDR')
        destinatarios = request.data.get('destinatarios', [])
        solicitud_txt = request.data.get('solicitudTexto', '')
        now           = timezone.now()

        if not destinatarios:
            return Response({'detail': 'Se requiere al menos un destinatario.'}, status=400)

        # Solo se puede turnar si el FUS está en Registrado o Turnado
        if fus.estatusParticular_id not in ('Registrado', 'Turnado'):
            return Response(
                {'detail': f'No se puede turnar un FUS en estado "{fus.estatusParticular_id}".'},
                status=400,
            )

        # Nombre del remitente (ROL1) para las notificaciones
        remitente_auth = CorreoAutorizado.objects.filter(email=user.email, activo=1).first()
        nombre_remitente = remitente_auth.nombre if remitente_auth else (user.first_name or user.email)

        for dest in destinatarios:
            dest_user = get_object_or_404(User, pk=dest['idDestinatario'])
            if _rol(dest_user) != 'ROL2':
                return Response({'detail': 'Destinatario inválido: debe ser un usuario ROL2 activo.'}, status=400)
            medio     = get_object_or_404(MedioRecepcion, pk=dest['idMedio'])

            ya_turnado = Turnado.objects.filter(idFus=fus, idDestinatario=dest_user, activo=1).exclude(estatusTitular_id='Concluido').exists()
            if ya_turnado:
                continue  # ya tiene un turnado activo, no duplicar

            Turnado.objects.create(
                idFus=fus,
                idRemitente=user,
                idDestinatario=dest_user,
                idMedio=medio,
                solicitudTexto=solicitud_txt,
                fechaHoraTurnado=now,
                idUsuarioRegistra=user.id,
            )
            _notif = Notificacion.objects.create(
                idDestinatario=dest_user,
                fusFolio=fus.folio,
                tipoEvento='TURNADO',
                mensaje=f"{nombre_remitente} te ha turnado el FUS {fus.folio}.",
            )
            _push_notificacion(_notif)

        estado_ant = fus.estatusParticular_id
        fus.estatusParticular_id = 'Turnado'
        fus.idUsuarioModifica = user.id
        fus.save()

        _log(usuario=user.email, rol=rol, accion='TURNAR_FUS',
             ip=ip, folio=fus.folio, estado_ant=estado_ant, estado_nuevo='Turnado')

        return Response({'detail': 'FUS turnado correctamente.'})


# ── Actividad de un FUS (ROL1 solo lectura) ──────────────────────────────────

class FUSActividadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        if _rol(request.user) != 'ROL1':
            return Response({'detail': 'No autorizado.'}, status=403)

        fus = get_object_or_404(FUS, pk=pk, activo=1, idSolicitanteInterno=request.user)
        turnados = Turnado.objects.filter(
            idFus=fus, activo=1
        ).select_related(
            'idDestinatario', 'idRemitente', 'idMedio',
        ).prefetch_related(
            'seguimientos',
        ).order_by('fechaHoraTurnado')
        return Response(TurnadoActividadSerializer(turnados, many=True).data)


class FUSDetalleAuditoriaView(APIView):
    """Detalle de auditoría de un FUS por folio, para el modal de Bitácora (ROL1)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, folio):
        if _rol(request.user) != 'ROL1':
            return Response({'detail': 'No autorizado.'}, status=403)

        fus = get_object_or_404(
            FUS.objects.select_related('idSolicitanteInterno', 'idMedioRecepcion', 'estatusParticular'),
            folio=folio, activo=1,
        )

        turnados = Turnado.objects.filter(idFus=fus).select_related(
            'idDestinatario'
        ).prefetch_related('seguimientos').order_by('fechaHoraTurnado')

        seguimientos = []
        estatus_titular = None
        for t in turnados:
            estatus_titular = t.estatusTitular_id
            autor = (t.idDestinatario.first_name or t.idDestinatario.email) if t.idDestinatario else None
            for s in t.seguimientos.all():
                if not s.activo:
                    continue
                seguimientos.append({
                    'fecha': s.fechaActividad,
                    'autor': autor,
                    'texto': s.descripcionActividad,
                })

        sol = fus.idSolicitanteInterno
        return Response({
            'folio': fus.folio,
            'descripcion': fus.descripcion,
            'contexto': fus.contexto,
            'medioRecepcion': fus.idMedioRecepcion.nombreMedio if fus.idMedioRecepcion else None,
            'prioridad': fus.prioridad,
            'criterios': fus.criterios,
            'nombreExterno': fus.nombreExterno,
            'telefonoExterno': fus.telefonoExterno,
            'correoExterno': fus.correoExterno,
            'estatusParticular': fus.estatusParticular_id,
            'estatusTitular': estatus_titular,
            'fechaRegistro': fus.fechaRegistro,
            'idSolicitanteInterno': {
                'nombre': resolver_nombre(sol) if sol else None,
                'email': sol.email if sol else None,
            },
            'evidencias': [{'nombreArchivo': e.nombreArchivo} for e in fus.evidencias.filter(activo=1)],
            'seguimientos': seguimientos,
        })


# ── Turnados (ROL2) ──────────────────────────────────────────────────────────

class MisTurnadosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Turnado.objects.filter(
            idDestinatario=request.user, activo=1
        ).select_related(
            'idFus', 'idFus__idSolicitanteInterno', 'idFus__idMedioRecepcion',
            'idRemitente', 'idMedio',
        )

        estatus = request.query_params.get('estatusTitular')
        search  = request.query_params.get('search')
        if estatus:
            qs = qs.filter(estatusTitular_id=estatus)
        if search:
            emails_nombre = list(CorreoAutorizado.objects.filter(nombre__icontains=search).values_list('email', flat=True))
            qs = qs.filter(
                Q(idFus__folio__icontains=search) |
                Q(idFus__descripcion__icontains=search) |
                Q(idFus__contexto__icontains=search) |
                Q(idFus__medioEspecificacion__icontains=search) |
                Q(idFus__criterios__icontains=search) |
                Q(idFus__nombreExterno__icontains=search) |
                Q(idFus__telefonoExterno__icontains=search) |
                Q(idFus__correoExterno__icontains=search) |
                Q(idFus__idMedioRecepcion__nombreMedio__icontains=search) |
                Q(idFus__idSolicitanteInterno__email__icontains=search) |
                Q(idFus__idSolicitanteInterno__email__in=emails_nombre) |
                Q(idFus__evidencias__nombreArchivo__icontains=search) |
                Q(idFus__evidencias__comentarios__icontains=search) |
                Q(solicitudTexto__icontains=search) |
                Q(idMedio__nombreMedio__icontains=search) |
                Q(idRemitente__email__icontains=search) |
                Q(idRemitente__email__in=emails_nombre) |
                Q(idDestinatario__email__icontains=search) |
                Q(idDestinatario__email__in=emails_nombre) |
                Q(seguimientos__descripcionActividad__icontains=search) |
                Q(seguimientos__accionTexto__icontains=search)
            ).distinct()

        qs = qs.order_by('-fechaRegistro')

        # Paginación
        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(100, max(1, int(request.query_params.get('page_size', 30))))
        except (ValueError, TypeError):
            page, page_size = 1, 30

        total  = qs.count()
        offset = (page - 1) * page_size
        data   = TurnadoSerializer(qs[offset: offset + page_size], many=True).data
        return Response({'total': total, 'page': page, 'page_size': page_size, 'results': data})


class ConcluirTurnadoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        turnado = get_object_or_404(Turnado, pk=pk, activo=1, idDestinatario=request.user)

        if turnado.estatusTitular_id == 'Concluido':
            return Response({'detail': 'El asunto ya está concluido.'}, status=400)

        # Condición obligatoria: debe existir al menos una respuesta registrada
        tiene_respuestas = Seguimiento.objects.filter(idTurnado=turnado, activo=1).exists()
        if not tiene_respuestas:
            return Response(
                {'detail': 'Debes registrar al menos una respuesta de seguimiento antes de concluir el asunto.'},
                status=400,
            )

        user    = request.user
        ip      = request.META.get('REMOTE_ADDR')
        rol     = _rol(user)
        est_ant = turnado.estatusTitular_id

        turnado.estatusTitular_id    = 'Concluido'
        turnado.idUsuarioModifica = user.id
        turnado.save()

        _log(usuario=user.email, rol=rol, accion='CONCLUSION_FUS',
             ip=ip, folio=turnado.idFus.folio,
             estado_ant=est_ant, estado_nuevo='Concluido')

        # Si TODOS los turnados activos del FUS están concluidos → FUS pasa a Concluido
        fus        = turnado.idFus
        pendientes = fus.turnados.filter(activo=1).exclude(estatusTitular_id='Concluido').count()
        if pendientes == 0:
            est_ant_fus              = fus.estatusParticular_id
            fus.estatusParticular_id = 'Concluido'
            fus.fechaConclusion   = timezone.now()
            fus.idUsuarioModifica = user.id
            fus.save()

            concluye_auth = CorreoAutorizado.objects.filter(email=user.email, activo=1).first()
            nombre_concluye = concluye_auth.nombre if concluye_auth else (user.first_name or user.email)
            _notif = Notificacion.objects.create(
                idDestinatario=fus.idSolicitanteInterno,
                fusFolio=fus.folio,
                tipoEvento='CONCLUIDO',
                mensaje=f"{nombre_concluye} ha concluido el FUS {fus.folio}.",
            )
            _push_notificacion(_notif)

            _log(usuario=user.email, rol=rol, accion='ASIGNACION_ESTADO',
                 ip=ip, folio=fus.folio,
                 estado_ant=est_ant_fus, estado_nuevo='Concluido')

        return Response({'detail': 'Asunto concluido correctamente.'})


# ── Seguimientos ─────────────────────────────────────────────────────────────

class SeguimientoListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, turnado_id):
        turnado = get_object_or_404(Turnado, pk=turnado_id, activo=1)
        if turnado.idDestinatario_id != request.user.id:
            return Response({'detail': 'No autorizado.'}, status=403)
        qs      = Seguimiento.objects.filter(idTurnado=turnado, activo=1).order_by('fechaActividad')
        return Response(SeguimientoSerializer(qs, many=True).data)

    def post(self, request, turnado_id):
        turnado = get_object_or_404(Turnado, pk=turnado_id, activo=1)
        if turnado.idDestinatario_id != request.user.id:
            return Response({'detail': 'No autorizado.'}, status=403)

        # Guard: bloquear si el asunto ya está concluido
        if turnado.estatusTitular_id == 'Concluido':
            return Response(
                {'detail': 'No se pueden agregar respuestas a un asunto ya concluido.'},
                status=400,
            )

        user = request.user
        ip   = request.META.get('REMOTE_ADDR')
        rol  = _rol(user)

        seg = Seguimiento.objects.create(
            idTurnado=turnado,
            fechaActividad=request.data.get('fechaActividad') or None,
            descripcionActividad=request.data.get('descripcionActividad', ''),
            accionTexto=request.data.get('accionTexto') or None,
            idUsuarioRegistra=user.id,
        )

        # Nombre del titular (ROL2) para las notificaciones
        titular_auth = CorreoAutorizado.objects.filter(email=user.email, activo=1).first()
        nombre_titular = titular_auth.nombre if titular_auth else (user.first_name or user.email)

        fus = turnado.idFus

        # Transición Recibido → En_seguimiento (primera respuesta del titular)
        if turnado.estatusTitular_id == 'Recibido':
            turnado.estatusTitular_id = 'En_seguimiento'
            turnado.idUsuarioModifica = user.id
            turnado.save()

            # Transición FUS: Turnado → Atendido (cuando al menos un titular responde)
            if fus.estatusParticular_id == 'Turnado':
                fus.estatusParticular_id = 'Atendido'
                fus.idUsuarioModifica = user.id
                fus.save()

                _notif = Notificacion.objects.create(
                    idDestinatario=fus.idSolicitanteInterno,
                    fusFolio=fus.folio,
                    tipoEvento='CAMBIO_ESTADO',
                    mensaje=f"{nombre_titular} comenzó a atender el FUS {fus.folio}.",
                )
                _push_notificacion(_notif)

                _log(usuario=user.email, rol=rol, accion='ASIGNACION_ESTADO',
                     ip=ip, folio=fus.folio,
                     estado_ant='Turnado', estado_nuevo='Atendido',
                     obs=f'Primera respuesta registrada por {nombre_titular}')
        else:
            # Seguimientos posteriores — notificar al ROL1 cada nueva respuesta
            resumen = seg.descripcionActividad[:80]
            if len(seg.descripcionActividad) > 80:
                resumen += '…'
            _notif = Notificacion.objects.create(
                idDestinatario=fus.idSolicitanteInterno,
                fusFolio=fus.folio,
                tipoEvento='RESPUESTA',
                mensaje=f"{nombre_titular} registró una nueva respuesta en el FUS {fus.folio}: \"{resumen}\"",
            )
            _push_notificacion(_notif)

        _log(usuario=user.email, rol=rol, accion='REGISTRO_RESPUESTA',
             ip=ip, folio=fus.folio)

        return Response(SeguimientoSerializer(seg).data, status=status.HTTP_201_CREATED)


class SeguimientoDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        seg = get_object_or_404(Seguimiento, pk=pk, activo=1)
        if seg.idTurnado.idDestinatario_id != request.user.id:
            return Response({'detail': 'No autorizado.'}, status=403)

        # Guard: bloquear si el asunto está concluido
        if seg.idTurnado.estatusTitular_id == 'Concluido':
            return Response(
                {'detail': 'No se pueden eliminar respuestas de un asunto ya concluido.'},
                status=400,
            )

        seg.activo = 0
        seg.save()
        return Response(status=status.HTTP_204_NO_CONTENT)




# ── Notificaciones ────────────────────────────────────────────────────────────

class NotificacionListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Notificacion.objects.filter(
            idDestinatario=request.user
        ).order_by('-fechaGeneracion')

        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(100, max(1, int(request.query_params.get('page_size', 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50

        total  = qs.count()
        offset = (page - 1) * page_size
        data   = NotificacionSerializer(qs[offset: offset + page_size], many=True).data
        return Response({'total': total, 'page': page, 'page_size': page_size, 'results': data})


class NotificacionMarcarLeidaView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        notif = get_object_or_404(Notificacion, pk=pk, idDestinatario=request.user)
        notif.leida = 1
        notif.fechaLectura = timezone.now()
        notif.save()
        return Response(NotificacionSerializer(notif).data)


class NotificacionMarcarTodasView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        Notificacion.objects.filter(
            idDestinatario=request.user, leida=0
        ).update(leida=1, fechaLectura=timezone.now())
        return Response({'detail': 'Todas marcadas como leídas.'})


# ── Bitácora ──────────────────────────────────────────────────────────────────

ROL1_ACCIONES = ['REGISTRO_RESPUESTA', 'CONCLUSION_FUS', 'ASIGNACION_ESTADO']
ROL2_ACCIONES = ['CONCLUSION_FUS', 'REGISTRO_RESPUESTA', 'REGISTRO_ACCION']


BITACORA_COLS_VALIDAS  = ['folio', 'nombre', 'usuario', 'fecha', 'accion', 'estado_ant', 'estado_nuevo', 'observaciones']
BITACORA_COLS_DEFAULT  = ['folio', 'nombre', 'usuario', 'fecha', 'accion']


def _parse_columnas_bitacora(request):
    raw = request.query_params.get('columnas')
    if not raw:
        return BITACORA_COLS_DEFAULT
    cols = [c.strip() for c in raw.split(',') if c.strip() in BITACORA_COLS_VALIDAS]
    return cols or BITACORA_COLS_DEFAULT


BITACORA_TITULO_PRINCIPAL = 'Sistema de control de solicitudes'
ANAM_LOGO_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    'frontend', 'src', 'assets', 'Logos_P_Hacienda_ANAM.png',
)


def _resolver_unidad_administrativa(user):
    autorizado = CorreoAutorizado.objects.select_related('unidadAdministrativa').filter(email=user.email).first()
    if autorizado and autorizado.unidadAdministrativa_id:
        return autorizado.unidadAdministrativa.unidadAdministrativa
    return 'Sin unidad asignada'


def _metadata_generacion():
    ahora = timezone.localtime(timezone.now())
    return f'Ciudad de México, {ahora.strftime("%d/%m/%Y")} a las {ahora.strftime("%H:%M")} h'


def _bitacora_base_qs(request):
    """Devuelve el queryset de bitácora ya filtrado por rol del usuario.
    Solo se incluyen registros ligados a un FUS (se excluyen eventos de cuenta
    como inicio/cierre de sesión o restablecimiento de contraseña)."""
    rol = _rol(request.user)
    if rol == 'ROL1':
        qs = Bitacora.objects.all()
    elif rol == 'ROL2':
        qs = Bitacora.objects.filter(
            usuario=request.user.email, accion__in=ROL2_ACCIONES
        )
    else:
        return Bitacora.objects.none()
    return qs.exclude(fusFolio__isnull=True).exclude(fusFolio='')


def _parse_fecha_local(fecha_str, fin_de_dia=False):
    """Convierte 'YYYY-MM-DD' (hora local del servidor) a datetime aware en UTC.
    Se evita el lookup `__date` de Django porque requiere CONVERT_TZ en MySQL,
    lo cual falla silenciosamente (devuelve NULL) si el servidor no tiene
    cargadas las tablas de zonas horarias, vaciando cualquier filtro de fecha."""
    try:
        y, m, d = (int(p) for p in fecha_str.split('-'))
    except (ValueError, AttributeError):
        return None
    naive = datetime.datetime(y, m, d, 23, 59, 59, 999999) if fin_de_dia else datetime.datetime(y, m, d, 0, 0, 0)
    return timezone.make_aware(naive, timezone.get_current_timezone())


def _aplicar_filtros_bitacora(qs, params, rol):
    usuario     = params.get('usuario')
    accion      = params.get('accion')
    folio       = params.get('folio')
    nombre      = params.get('nombre')
    estatus_fus = params.get('estatus_fus')
    fecha_desde = params.get('fecha_desde')
    fecha_hasta = params.get('fecha_hasta')

    q = params.get('q')
    if q:
        if rol == 'ROL1':
            emails_nombre = CorreoAutorizado.objects.filter(nombre__icontains=q).values_list('email', flat=True)
            qs = qs.filter(Q(fusFolio__icontains=q) | Q(usuario__icontains=q) | Q(usuario__in=list(emails_nombre)))
        else:
            qs = qs.filter(fusFolio__icontains=q)

    if usuario and rol == 'ROL1': qs = qs.filter(usuario__icontains=usuario)
    if accion:      qs = qs.filter(accion=accion)
    if folio:       qs = qs.filter(fusFolio__icontains=folio)

    if nombre and rol == 'ROL1':
        emails = CorreoAutorizado.objects.filter(nombre__icontains=nombre).values_list('email', flat=True)
        qs = qs.filter(usuario__in=list(emails))

    if estatus_fus:
        folios = FUS.objects.filter(estatusParticular_id=estatus_fus).values_list('folio', flat=True)
        qs = qs.filter(fusFolio__in=list(folios))

    dt_desde = _parse_fecha_local(fecha_desde) if fecha_desde else None
    if dt_desde: qs = qs.filter(fechaHora__gte=dt_desde)

    dt_hasta = _parse_fecha_local(fecha_hasta, fin_de_dia=True) if fecha_hasta else None
    if dt_hasta: qs = qs.filter(fechaHora__lte=dt_hasta)

    return qs


class BitacoraListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rol = _rol(request.user)
        ordering = request.query_params.get('ordering')
        qs = _bitacora_base_qs(request)
        if not ordering:
            qs = qs.order_by('-fechaHora')
        qs  = _aplicar_filtros_bitacora(qs, request.query_params, rol)

        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(100, max(1, int(request.query_params.get('page_size', 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50

        ORDERING_MAP = {
            'folio': 'fusFolio', 'fecha': 'fechaHora', 'accion': 'accion',
        }
        if ordering:
            campo = ordering.lstrip('-')
            campo_real = ORDERING_MAP.get(campo)
            if campo_real:
                qs = qs.order_by(f"{'-' if ordering.startswith('-') else ''}{campo_real}")

        total  = qs.count()
        offset = (page - 1) * page_size

        from .serializers import BitacoraSerializer
        pagina = list(qs[offset: offset + page_size])
        nombres_map = dict(
            CorreoAutorizado.objects.filter(
                email__in={r.usuario for r in pagina}
            ).values_list('email', 'nombre')
        )
        data = BitacoraSerializer(pagina, many=True, context={'nombres_map': nombres_map}).data
        return Response({
            'total': total, 'page': page, 'page_size': page_size,
            'results': data, 'rol': rol,
        })


# ── Exportar Bitácora ────────────────────────────────────────────────────────

class ExportarBitacoraExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage

        rol = _rol(request.user)
        qs  = _bitacora_base_qs(request).order_by('-fechaHora')
        qs  = _aplicar_filtros_bitacora(qs, request.query_params, rol)
        columnas = _parse_columnas_bitacora(request)

        ACCION_LABELS = {
            'REGISTRO_FUS': 'Registro FUS', 'TURNAR_FUS': 'Turnar FUS',
            'ASIGNACION_ESTADO': 'Cambio de estado', 'REGISTRO_RESPUESTA': 'Registro respuesta',
            'REGISTRO_ACCION': 'Registro acción', 'CONCLUSION_FUS': 'Conclusión FUS',
            'INICIO_SESION': 'Inicio sesión', 'CIERRE_SESION': 'Cierre sesión',
            'RESTABLECER_CONTRASENA': 'Restablecer contraseña',
            'ELIMINACION': 'Eliminación',
        }
        nombres_map = dict(
            CorreoAutorizado.objects.filter(
                email__in=qs.values_list('usuario', flat=True).distinct()
            ).values_list('email', 'nombre')
        )
        col_defs = {
            'folio':         ('Folio',           lambda r: r.fusFolio or ''),
            'nombre':        ('Nombre',          lambda r: nombres_map.get(r.usuario, '')),
            'usuario':       ('Usuario',         lambda r: r.usuario),
            'fecha':         ('Fecha y hora',    lambda r: r.fechaHora.strftime('%d/%m/%Y %H:%M:%S') if r.fechaHora else ''),
            'accion':        ('Acción',          lambda r: ACCION_LABELS.get(r.accion, r.accion)),
            'estado_ant':    ('Estado anterior', lambda r: r.estadoAnterior or ''),
            'estado_nuevo':  ('Estado nuevo',    lambda r: r.estadoNuevo or ''),
            'observaciones': ('Observaciones',   lambda r: r.observaciones or ''),
        }
        headers = [col_defs[c][0] for c in columnas]
        n_cols  = len(headers)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bitácora'

        unidad_admin    = _resolver_unidad_administrativa(request.user)
        metadata_linea  = _metadata_generacion()

        # Encabezado institucional
        ws.append(['AGENCIA NACIONAL DE ADUANAS DE MÉXICO'])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Título principal
        ws.append([BITACORA_TITULO_PRINCIPAL.upper()])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Subtítulo dinámico — unidad/aduana asignada al usuario
        ws.append([unidad_admin.upper()])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(bold=True, size=11, color='1F5647')
        cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.append([])

        # Metadatos — fecha/hora de generación (derecha)
        ws.append([metadata_linea])
        fila = ws.max_row
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
        cell = ws.cell(row=fila, column=1)
        cell.font = Font(italic=True, size=9, color='595959')
        cell.alignment = Alignment(horizontal='right')

        # Logo institucional (izquierda)
        logo_row = ws.max_row + 1
        for _ in range(5):
            ws.append([])
        if os.path.exists(ANAM_LOGO_PATH):
            img = XLImage(ANAM_LOGO_PATH)
            img.width, img.height = 460, 460 * (150/889)
            ws.add_image(img, f'A{logo_row}')

        ws.append([])

        ws.append(headers)
        header_row = ws.max_row
        fill   = PatternFill('solid', fgColor='FFFF00')
        thin   = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[header_row]:
            cell.fill = fill
            cell.font = Font(bold=True, color='000000', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        for r in qs:
            ws.append([col_defs[c][1](r) for c in columnas])
            for cell in ws[ws.max_row]:
                cell.border = border

        for i in range(1, n_cols + 1):
            col_letter = get_column_letter(i)
            max_len = max(
                (len(str(ws.cell(row=r, column=i).value or '')) for r in range(header_row, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        ws.freeze_panes = f'A{header_row + 1}'

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename="bitacora.xlsx"'
        wb.save(resp)
        return resp


class ExportarBitacoraPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io

        rol = _rol(request.user)
        qs  = _bitacora_base_qs(request).order_by('-fechaHora')
        qs  = _aplicar_filtros_bitacora(qs, request.query_params, rol)
        columnas = _parse_columnas_bitacora(request)

        ACCION_LABELS = {
            'REGISTRO_FUS': 'Registro FUS', 'TURNAR_FUS': 'Turnar FUS',
            'ASIGNACION_ESTADO': 'Cambio de estado', 'REGISTRO_RESPUESTA': 'Registro respuesta',
            'REGISTRO_ACCION': 'Registro acción', 'CONCLUSION_FUS': 'Conclusión FUS',
            'INICIO_SESION': 'Inicio sesión', 'CIERRE_SESION': 'Cierre sesión',
            'RESTABLECER_CONTRASENA': 'Restablecer contraseña',
            'ELIMINACION': 'Eliminación',
        }
        nombres_map = dict(
            CorreoAutorizado.objects.filter(
                email__in=qs.values_list('usuario', flat=True).distinct()
            ).values_list('email', 'nombre')
        )

        styles = getSampleStyleSheet()
        cell_style  = ParagraphStyle('cell', parent=styles['Normal'], fontSize=7.5, leading=10)

        col_defs = {
            'folio':         ('Folio',         lambda r: r.fusFolio or '—'),
            'nombre':        ('Nombre',        lambda r: nombres_map.get(r.usuario, '—')),
            'usuario':       ('Usuario',       lambda r: r.usuario),
            'fecha':         ('Fecha y hora',  lambda r: r.fechaHora.strftime('%d/%m/%Y %H:%M') if r.fechaHora else '—'),
            'accion':        ('Acción',        lambda r: ACCION_LABELS.get(r.accion, r.accion)),
            'estado_ant':    ('Estado ant.',   lambda r: r.estadoAnterior or '—'),
            'estado_nuevo':  ('Estado nuevo',  lambda r: r.estadoNuevo or '—'),
            'observaciones': ('Observaciones', lambda r: r.observaciones or '—'),
        }
        headers = [col_defs[c][0] for c in columnas]

        unidad_admin   = _resolver_unidad_administrativa(request.user)
        metadata_linea = _metadata_generacion()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1.2*cm, rightMargin=1.2*cm,
                                topMargin=4.4*cm, bottomMargin=1.2*cm)

        ancho_total = landscape(A4)[0] - 2.4*cm

        def _encabezado_bitacora(canvas_, doc_):
            canvas_.saveState()
            page_w, page_h = landscape(A4)
            centro = page_w / 2
            derecha = page_w - 1.2*cm
            y = page_h - 1.1*cm

            canvas_.setFont('Helvetica-Bold', 11)
            canvas_.setFillColor(colors.black)
            canvas_.drawCentredString(centro, y, 'AGENCIA NACIONAL DE ADUANAS DE MÉXICO')
            y -= 0.42*cm
            canvas_.drawCentredString(centro, y, BITACORA_TITULO_PRINCIPAL.upper())
            y -= 0.42*cm
            canvas_.setFillColor(colors.HexColor('#1F5647'))
            canvas_.drawCentredString(centro, y, unidad_admin.upper())
            canvas_.setFillColor(colors.black)

            y -= 0.5*cm
            canvas_.setFont('Helvetica-Oblique', 7)
            canvas_.setFillColor(colors.HexColor('#595959'))
            canvas_.drawRightString(derecha, y, metadata_linea)
            canvas_.setFillColor(colors.black)

            if os.path.exists(ANAM_LOGO_PATH):
                logo_w = 8*cm
                logo_h = logo_w * (150/889)
                canvas_.drawImage(ANAM_LOGO_PATH, 1.2*cm, y - logo_h - 0.2*cm,
                                  width=logo_w, height=logo_h, mask='auto', preserveAspectRatio=True)
            canvas_.restoreState()

        elements = []

        data = [headers]
        for r in qs:
            data.append([Paragraph(str(col_defs[c][1](r)), cell_style) for c in columnas])

        col_width = ancho_total / len(columnas)
        t = Table(data, colWidths=[col_width] * len(columnas), repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFFF00')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.black),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 8.5),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)

        doc.build(elements, onFirstPage=_encabezado_bitacora, onLaterPages=_encabezado_bitacora)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="bitacora.pdf"'
        return resp


# ── Exportar FUS ──────────────────────────────────────────────────────────────

def _fus_queryset(request):
    if _rol(request.user) != 'ROL1':
        raise PermissionDenied('No autorizado.')
    qs = FUS.objects.filter(activo=1).select_related(
        'idSolicitanteInterno', 'idMedioRecepcion'
    ).order_by('-fechaRegistro')
    estatus = request.query_params.get('estatusParticular')
    search  = request.query_params.get('search')
    if estatus: qs = qs.filter(estatusParticular_id=estatus)
    if search:
        emails_nombre = list(CorreoAutorizado.objects.filter(nombre__icontains=search).values_list('email', flat=True))
        qs = qs.filter(
            Q(folio__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(contexto__icontains=search) |
            Q(medioEspecificacion__icontains=search) |
            Q(criterios__icontains=search) |
            Q(nombreExterno__icontains=search) |
            Q(telefonoExterno__icontains=search) |
            Q(correoExterno__icontains=search) |
            Q(idMedioRecepcion__nombreMedio__icontains=search) |
            Q(idSolicitanteInterno__email__icontains=search) |
            Q(idSolicitanteInterno__email__in=emails_nombre) |
            Q(evidencias__nombreArchivo__icontains=search) |
            Q(evidencias__comentarios__icontains=search) |
            Q(turnados__solicitudTexto__icontains=search) |
            Q(turnados__idMedio__nombreMedio__icontains=search) |
            Q(turnados__idRemitente__email__icontains=search) |
            Q(turnados__idRemitente__email__in=emails_nombre) |
            Q(turnados__idDestinatario__email__icontains=search) |
            Q(turnados__idDestinatario__email__in=emails_nombre) |
            Q(turnados__seguimientos__descripcionActividad__icontains=search) |
            Q(turnados__seguimientos__accionTexto__icontains=search)
        ).distinct()
    return qs


class ExportarFUSExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse

        qs = _fus_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Solicitudes FUS'

        headers = ['Folio', 'Fecha', 'Solicitante', 'Medio', 'Prioridad',
                   'Estatus', 'Descripción', 'Contexto']
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill('solid', fgColor='9F2241')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        fmt = lambda d: d.strftime('%d/%m/%Y %H:%M') if d else ''
        fmt_date = lambda d: d.strftime('%d/%m/%Y') if d else ''

        for fus in qs:
            sol = fus.idSolicitanteInterno
            nombre_sol = f"{sol.first_name} {sol.last_name}".strip() if sol else ''
            ws.append([
                fus.folio,
                fmt(fus.fechaHora),
                nombre_sol or (sol.email if sol else ''),
                fus.idMedioRecepcion.nombreMedio if fus.idMedioRecepcion else '',
                fus.prioridad or '',
                fus.estatusParticular_id,
                fus.descripcion,
                fus.contexto or '',
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="solicitudes_fus.xlsx"'
        wb.save(response)
        return response


class ExportarFUSPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io

        qs = _fus_queryset(request)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                     fontSize=16, textColor=colors.HexColor('#9F2241'),
                                     spaceAfter=6)
        cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)

        elements = [
            Paragraph('Reporte de Solicitudes FUS — ANAM', title_style),
            Spacer(1, 0.3*cm),
        ]

        fmt = lambda d: d.strftime('%d/%m/%Y %H:%M') if d else '—'
        fmt_date = lambda d: d.strftime('%d/%m/%Y') if d else '—'

        data = [['Folio', 'Fecha', 'Solicitante', 'Prioridad', 'Estatus', 'Descripción']]
        for fus in qs:
            sol = fus.idSolicitanteInterno
            nombre_sol = f"{sol.first_name} {sol.last_name}".strip() if sol else (sol.email if sol else '—')
            desc = fus.descripcion[:120] + ('…' if len(fus.descripcion) > 120 else '')
            data.append([
                Paragraph(fus.folio, cell_style),
                Paragraph(fmt(fus.fechaHora), cell_style),
                Paragraph(nombre_sol, cell_style),
                Paragraph(fus.prioridad or '—', cell_style),
                Paragraph(fus.estatusParticular_id, cell_style),
                Paragraph(desc, cell_style),
            ])

        col_widths = [4*cm, 3.5*cm, 3.5*cm, 2*cm, 2.5*cm, None]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F4F6')]),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#DDD0D5')),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="solicitudes_fus.pdf"'
        return response


# ── Descargar FUS individual (PDF) ────────────────────────────────────────────

class DescargarFUSPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, folio):
        from django.conf import settings
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, Image as RLImage, PageBreak, KeepTogether,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io

        try:
            fus = FUS.objects.select_related(
                'idSolicitanteInterno', 'idMedioRecepcion', 'estatusParticular'
            ).prefetch_related(
                'evidencias', 'turnados__idDestinatario', 'turnados__idMedio', 'turnados__seguimientos'
            ).get(folio=folio, activo=1)
        except FUS.DoesNotExist:
            from rest_framework.response import Response
            return Response({'detail': 'FUS no encontrado.'}, status=404)

        incluir_imagenes = request.query_params.get('imagenes') == '1'

        evidencias = [e for e in fus.evidencias.all() if e.activo]
        turnados = [t for t in fus.turnados.all() if t.activo]

        LETTERHEAD_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'membretada.png')

        def _membrete(canvas_, doc_):
            canvas_.saveState()
            if os.path.exists(LETTERHEAD_PATH):
                canvas_.drawImage(
                    LETTERHEAD_PATH, 0, 0,
                    width=letter[0], height=letter[1],
                    mask='auto', preserveAspectRatio=False,
                )
            canvas_.restoreState()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=3.3*cm, bottomMargin=2.8*cm)
        W = letter[0] - 4*cm  # ancho útil

        VERDE    = colors.black
        AMARILLO = colors.HexColor('#FFFF00')
        CLARO    = colors.white
        BORDE    = colors.black

        styles = getSampleStyleSheet()

        st_titulo = ParagraphStyle('titulo', fontName='Helvetica-Bold', fontSize=16,
                                   textColor=colors.black, spaceAfter=2)
        st_folio  = ParagraphStyle('folio',  fontName='Helvetica-Bold', fontSize=11,
                                   textColor=colors.black, spaceBefore=6, spaceAfter=8)
        st_sec    = ParagraphStyle('sec',    fontName='Helvetica-Bold', fontSize=9,
                                   textColor=colors.black, spaceAfter=0)
        st_lbl    = ParagraphStyle('lbl',    fontName='Helvetica-Bold', fontSize=8,
                                   textColor=colors.black)
        st_val    = ParagraphStyle('val',    fontName='Helvetica',      fontSize=8,
                                   textColor=colors.black, leading=11)

        fmt = lambda d: d.strftime('%d/%m/%Y %H:%M') if d else '—'

        sol = fus.idSolicitanteInterno
        nombre_sol = resolver_nombre(sol) if sol else '—'

        elements = []

        # ── Encabezado ──
        elements.append(Paragraph('FORMATO ÚNICO DE SOLICITUD', st_titulo))
        elements.append(Paragraph(
            f'Folio: {fus.folio} &nbsp;|&nbsp; Estatus: {fus.estatusParticular_id}',
            st_folio,
        ))
        elements.append(HRFlowable(width='100%', thickness=2, color=VERDE, spaceAfter=10))

        def seccion(titulo):
            t = Table([[Paragraph(titulo, st_sec)]], colWidths=[W])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), AMARILLO),
                ('TOPPADDING',    (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ]))
            return t

        def fila(lbl, val):
            return [Paragraph(lbl, st_lbl), Paragraph(str(val) if val else '—', st_val)]

        # ── Datos generales ──
        datos = [
            fila('Fecha y hora',        fmt(fus.fechaHora)),
            fila('Medio de recepción',  fus.idMedioRecepcion.nombreMedio if fus.idMedioRecepcion else '—'),
            fila('Solicitante interno', nombre_sol),
        ]
        dt = Table(datos, colWidths=[4*cm, W - 4*cm])
        dt.setStyle(TableStyle([
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CLARO]),
            ('GRID',           (0,0), (-1,-1), 0.3, BORDE),
            ('TOPPADDING',     (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',  (0,0), (-1,-1), 4),
            ('LEFTPADDING',    (0,0), (-1,-1), 6),
            ('RIGHTPADDING',   (0,0), (-1,-1), 6),
        ]))
        elements.append(KeepTogether([seccion('DATOS GENERALES'), Spacer(1, 4), dt]))
        elements.append(Spacer(1, 8))

        # ── Descripción ──
        desc_data = [
            fila('Descripción', fus.descripcion),
            fila('Datos o antecedentes de contexto de la solicitud', fus.contexto or '—'),
        ]
        dt2 = Table(desc_data, colWidths=[4*cm, W - 4*cm])
        dt2.setStyle(TableStyle([
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CLARO]),
            ('GRID',           (0,0), (-1,-1), 0.3, BORDE),
            ('TOPPADDING',     (0,0), (-1,-1), 6),
            ('BOTTOMPADDING',  (0,0), (-1,-1), 6),
            ('LEFTPADDING',    (0,0), (-1,-1), 6),
            ('RIGHTPADDING',   (0,0), (-1,-1), 6),
            ('VALIGN',         (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(KeepTogether([seccion('DESCRIPCIÓN DE LA SOLICITUD'), Spacer(1, 4), dt2]))
        elements.append(Spacer(1, 8))

        # ── Solicitante externo ──
        if fus.nombreExterno or fus.correoExterno or fus.telefonoExterno:
            ext_data = [
                fila('Nombre',    fus.nombreExterno),
                fila('Correo',    fus.correoExterno),
                fila('Teléfono',  fus.telefonoExterno),
            ]
            dt3 = Table(ext_data, colWidths=[4*cm, W - 4*cm])
            dt3.setStyle(TableStyle([
                ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CLARO]),
                ('GRID',           (0,0), (-1,-1), 0.3, BORDE),
                ('TOPPADDING',     (0,0), (-1,-1), 4),
                ('BOTTOMPADDING',  (0,0), (-1,-1), 4),
                ('LEFTPADDING',    (0,0), (-1,-1), 6),
                ('RIGHTPADDING',   (0,0), (-1,-1), 6),
            ]))
            elements.append(KeepTogether([seccion('SOLICITANTE EXTERNO'), Spacer(1, 4), dt3]))
            elements.append(Spacer(1, 8))

        # ── Evidencia (solo nombres de archivo; las imágenes van al final si se solicitaron) ──
        if evidencias:
            ev_rows = []
            for ev in evidencias:
                texto = ev.nombreArchivo or '—'
                if ev.comentarios:
                    texto += f' — {ev.comentarios}'
                ev_rows.append([Paragraph(texto, st_val)])
            evt = Table(ev_rows, colWidths=[W])
            evt.setStyle(TableStyle([
                ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CLARO]),
                ('GRID',           (0,0), (-1,-1), 0.3, BORDE),
                ('TOPPADDING',     (0,0), (-1,-1), 5),
                ('BOTTOMPADDING',  (0,0), (-1,-1), 5),
                ('LEFTPADDING',    (0,0), (-1,-1), 8),
            ]))
            elements.append(KeepTogether([seccion('EVIDENCIA'), Spacer(1, 4), evt]))
        else:
            elements.append(KeepTogether([seccion('EVIDENCIA'), Spacer(1, 4), Paragraph('—', st_val)]))
        elements.append(Spacer(1, 8))

        # ── Prioridad ──
        prioridad_bloque = [seccion('PRIORIDAD'), Spacer(1, 4), Paragraph(f'<b>{fus.prioridad or "—"}</b>', st_val)]
        if fus.criterios:
            prioridad_bloque.append(Spacer(1, 2))
            for crit in [c.strip() for c in fus.criterios.split('|') if c.strip()]:
                prioridad_bloque.append(Paragraph(f'• {crit}', st_val))
        elements.append(KeepTogether(prioridad_bloque))
        elements.append(Spacer(1, 8))

        # ── Se turnó ──
        if turnados:
            for i, t in enumerate(turnados):
                dest_nombre = resolver_nombre(t.idDestinatario) if t.idDestinatario else '—'
                turno_data = [
                    fila('Nombre',             dest_nombre),
                    fila('Medio de envío',     t.idMedio.nombreMedio if t.idMedio else '—'),
                    fila('Fecha y hora',       fmt(t.fechaHoraTurnado)),
                    fila('Texto de la solicitud', t.solicitudTexto or '—'),
                ]
                tnt = Table(turno_data, colWidths=[4*cm, W - 4*cm])
                tnt.setStyle(TableStyle([
                    ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CLARO]),
                    ('GRID',           (0,0), (-1,-1), 0.3, BORDE),
                    ('TOPPADDING',     (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING',  (0,0), (-1,-1), 4),
                    ('LEFTPADDING',    (0,0), (-1,-1), 6),
                    ('RIGHTPADDING',   (0,0), (-1,-1), 6),
                    ('VALIGN',         (0,0), (-1,-1), 'TOP'),
                ]))
                bloque = [seccion('SE TURNÓ'), Spacer(1, 4), tnt] if i == 0 else [tnt]
                elements.append(KeepTogether(bloque))
                elements.append(Spacer(1, 6))

        # ── Respuesta y seguimiento ──
        turnados_con_seguimiento = [
            (t, [s for s in t.seguimientos.all() if s.activo]) for t in turnados
        ]
        turnados_con_seguimiento = [(t, segs) for t, segs in turnados_con_seguimiento if segs]

        if turnados:
            if turnados_con_seguimiento:
                for i, (t, segs) in enumerate(turnados_con_seguimiento):
                    dest_nombre = resolver_nombre(t.idDestinatario) if t.idDestinatario else '—'
                    seg_rows = []
                    for s in segs:
                        fecha_str = s.fechaActividad.strftime('%d/%m/%Y') if s.fechaActividad else '—'
                        texto = s.descripcionActividad or '—'
                        if s.accionTexto:
                            texto += f'<br/>→ {s.accionTexto}'
                        seg_rows.append([Paragraph(fecha_str, st_val), Paragraph(texto, st_val)])
                    segt = Table(seg_rows, colWidths=[3*cm, W - 3*cm])
                    segt.setStyle(TableStyle([
                        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, CLARO]),
                        ('GRID',           (0,0), (-1,-1), 0.3, BORDE),
                        ('TOPPADDING',     (0,0), (-1,-1), 4),
                        ('BOTTOMPADDING',  (0,0), (-1,-1), 4),
                        ('LEFTPADDING',    (0,0), (-1,-1), 6),
                        ('RIGHTPADDING',   (0,0), (-1,-1), 6),
                        ('VALIGN',         (0,0), (-1,-1), 'TOP'),
                    ]))
                    bloque = [Paragraph(dest_nombre, st_lbl), Spacer(1, 2), segt]
                    if i == 0:
                        bloque = [seccion('RESPUESTA Y SEGUIMIENTO'), Spacer(1, 4)] + bloque
                    elements.append(KeepTogether(bloque))
                    elements.append(Spacer(1, 8))
            else:
                elements.append(KeepTogether([
                    seccion('RESPUESTA Y SEGUIMIENTO'), Spacer(1, 4),
                    Paragraph('Pendiente de respuesta del titular.', st_val),
                ]))
                elements.append(Spacer(1, 8))

        # ── Pie ──
        elements.append(Spacer(1, 12))
        elements.append(HRFlowable(width='100%', thickness=0.5, color=BORDE))
        pie = ParagraphStyle('pie', fontName='Helvetica', fontSize=7,
                             textColor=colors.HexColor('#888888'), spaceBefore=4)
        from datetime import datetime, timezone
        now_str = datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')
        elements.append(Paragraph(
            f'Agencia Nacional de Aduanas de México — Sistema de Control de Solicitudes — Generado: {now_str}',
            pie
        ))

        # ── Anexo de imágenes de evidencia (hoja nueva, al final) ──
        if incluir_imagenes:
            imagenes = [e for e in evidencias if (e.tipoMime or '').startswith('image/')]
            rutas_validas = []
            for ev in imagenes:
                ruta_abs = os.path.join(settings.MEDIA_ROOT, ev.rutaArchivo or '')
                if ev.rutaArchivo and os.path.exists(ruta_abs):
                    rutas_validas.append((ev, ruta_abs))

            if rutas_validas:
                elements.append(PageBreak())
                elements.append(Paragraph('ANEXO — IMÁGENES DE EVIDENCIA', st_titulo))
                elements.append(HRFlowable(width='100%', thickness=2, color=VERDE, spaceAfter=12))
                max_w, max_h = W, 20*cm
                for ev, ruta_abs in rutas_validas:
                    elements.append(Paragraph(ev.nombreArchivo or '—', st_lbl))
                    elements.append(Spacer(1, 4))
                    try:
                        img = RLImage(ruta_abs)
                        ratio = min(max_w / img.imageWidth, max_h / img.imageHeight, 1)
                        img.drawWidth  = img.imageWidth * ratio
                        img.drawHeight = img.imageHeight * ratio
                        elements.append(img)
                    except Exception:
                        elements.append(Paragraph('(No se pudo cargar la imagen)', st_val))
                    elements.append(Spacer(1, 16))

        doc.build(elements, onFirstPage=_membrete, onLaterPages=_membrete)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="FUS_{folio}.pdf"'
        return resp
